import winapps
import subprocess
import socket
from openpyxl import Workbook  
import platform
import os
from openpyxl.styles import Font



# get each application with list_installed()

def get_serialnumber():
    Data = subprocess.check_output(["wmic", "bios", "get", "serialnumber"])
    serial_number = str(Data)
    removal_element = [r"\r", r"\n", "b'", "'"]
    array_length = len(removal_element)
    for i in range(array_length):
        serial_number = serial_number.replace(r"{}".format(removal_element[i]), '')
    return(serial_number.strip())

def model_number():
    Model_Data = subprocess.check_output(["wmic", "csproduct", "get", "name"])
    model_number = str(Model_Data)
    removal_element = [r"\r", r"\n", "b'", "'"]
    array_length = len(removal_element)
    for i in range(array_length):
        model_number = model_number.replace(r"{}".format(removal_element[i]), '')
    model_number = model_number.replace('Name', '')
    return(model_number.strip())

def host_name():
    Hostname=socket.gethostname()
    return(Hostname)


# f = open("data.txt", "w")
# for item in winapps.list_installed():
#     f.write("Application name:{} \t Version:{} \n".format(item.name,item.version))
# f.write("{} \n {}".format(get_serialnumber(),host_name()))
# f.close()

#Excel sheet
wb = Workbook()  
sheet = wb.active  

sheet.cell(row=1, column=1).value = "Serial Number" 
sheet.cell(row=1, column=2).value = "Host Name"  
sheet.cell(row=1, column=3).value = "Model Number"
sheet.cell(row=1, column=4).value = "Operating System Name"

for i in range(0,4):
    sheet.cell(row=1, column=i+1).font = Font(bold=True)



sheet.cell(row=2, column=1).value=get_serialnumber()
sheet.cell(row=2, column=2).value=host_name()
sheet.cell(row=2, column=3).value=model_number()
sheet.cell(row=2, column=4).value=platform.platform()


sheet.cell(row=4,column=1).value="Application Name"
sheet.cell(row=4,column=2).value="Version"

for i in range(0,2):
    sheet.cell(row=4, column=i+1).font = Font(bold=True)

count=0
for item in winapps.list_installed():
    sheet.cell(row=5+count,column=1).value=item.name
    sheet.cell(row=5+count,column=2).value=item.version
    count+=1


for i in range(0,4):
    sheet.column_dimensions[chr(65+i)].width = 30


wb.save("Extract.xlsx")  


